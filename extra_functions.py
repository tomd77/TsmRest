import os.path
import csv

# 3rd Party Modules
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle, Color


# Function to init Excel styles
def create_xls_styles(cur_workbook):

    # Title cell
    if 'title_style' not in cur_workbook.named_styles:
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(bold=True, size=12, color="043891")
        title_style.fill = PatternFill("solid", fgColor="DDDDDD")
        bd = Side(style='thin', color="000000")
        title_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        title_style.alignment = Alignment(vertical="center")
        cur_workbook.add_named_style(title_style)

    # Colored cells
    if 'green_cell' not in cur_workbook.named_styles:
        green_cell = NamedStyle(name="green_cell")
        green_cell.font = Font(bold=True)
        green_cell.fill = PatternFill("solid", fgColor="83CA86")
        cur_workbook.add_named_style(green_cell)

    if 'red_cell' not in cur_workbook.named_styles:
        red_cell = NamedStyle(name="red_cell")
        red_cell.font = Font(bold=True)
        red_cell.fill = PatternFill("solid", fgColor="EB6368")
        cur_workbook.add_named_style(red_cell)

    if 'dark_green_cell' not in cur_workbook.named_styles:
        dark_green_cell = NamedStyle(name="dark_green_cell")
        dark_green_cell.font = Font(bold=True)
        dark_green_cell.fill = PatternFill("solid", fgColor="2C5B2C")
        cur_workbook.add_named_style(dark_green_cell)

    if 'dark_red_cell' not in cur_workbook.named_styles:
        dark_red_cell = NamedStyle(name="dark_red_cell")
        dark_red_cell.font = Font(bold=True)
        dark_red_cell.fill = PatternFill("solid", fgColor="D24949")
        cur_workbook.add_named_style(dark_red_cell)

# End of function create_xls_styles


def read_excel_sheet(excel_file, excel_sheet):
    # EXPECTING sheet row 1 to contain 'column titles'

    if not os.path.exists(os.path.abspath(os.path.dirname(excel_file))):
        print("read_structured_sheet(): Directory not found")
        exit(1)

    if not os.path.exists(excel_file):
        print("read_structured_sheet(): Excel file not found")
        exit(1)

    wb = load_workbook(filename=excel_file)
    ws = wb[excel_sheet]

    # Get data from first row
    column_names = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):  # iter_rows returns tuple
        for cell in row:
            column_names.append(cell)

    # Get data from all rows after the first row
    excel_data = []
    for row in ws[2: ws.max_row]:  # not using iter_rows, need 'generator' to access col_idx
        cur_row = {}
        for cell in row:
            cur_row[column_names[cell.col_idx - 1]] = cell.value
        excel_data.append(cur_row)

    wb.close()

    return_data = {'hdr': column_names, 'items': excel_data}
    return return_data
    # End of function read_excel_sheet()


def read_csv(csv_file):
    csv_data = {}
    with open(csv_file, 'r') as f:
        reader = csv.DictReader(f)
        csv_data['items'] = list(reader)
        csv_data['hdr'] = reader.fieldnames

    return csv_data
