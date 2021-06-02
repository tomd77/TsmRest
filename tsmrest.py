import json
import ssl
import base64
from urllib.request import urlopen, Request
import urllib.error
import socket
import csv
import os.path
import datetime
from pprint import pprint

# 3rd Party Modules
from openpyxl import Workbook, load_workbook

# Project specific functions
from extra_functions import create_xls_styles


class TsmRest:

    request_timeout = 10  # HTTP request timeout (seconds)

    def __init__(self, oc_address, oc_port):
        self.oc_address = oc_address
        self.oc_port = oc_port
        self.tsm_servers = []
        self.tsm_command = None
        self.raw_result = []        # REST response - structured by IBM
        self.parsed_result = {}     # REST response - restructured with parse_raw_data()

    @property
    def base_url(self):
        return str("https://" + self.oc_address + ":" + self.oc_port + "/oc")

    def run_command(self, tsm_servers, tsm_user, tsm_pass, tsm_command):
        # Send TSM command via REST call to TSM OC web server
        # Nothing is returned to main. self.parsed_result and self.raw_result will be set.
        self.parsed_result = {}
        self.raw_result = []

        # Convert tsm_servers to list if only 1 TSM server is given as a string
        if isinstance(tsm_servers, str):
            tsm_servers = [tsm_servers]
        self.tsm_servers = tsm_servers

        print(f"\nExecuting your command on {len(tsm_servers)} TSM server(s)...")
        for tsm_server in tsm_servers:

            # Prepare HTTP request:

            # Credential string for HTTP header + Re-encode to ensure all chars are safe for transport
            cred_str = tsm_user + ":" + tsm_pass
            cred_encoded = base64.b64encode(cred_str.encode()).decode()

            request_header = {
                "OC-API-Version": "1.0",
                "Authorization": "Basic " + cred_encoded,
                "Accept": "application/json",
                "Content-type": "text/plain"
            }
            request_payload = tsm_command

            # Construct full URL
            full_url = self.base_url + '/api/cli/issueConfirmedCommand/' + tsm_server

            # Create HTTP request object
            my_http_request = Request(full_url, request_payload.encode(), request_header, method='POST')

            # Create empty SSL context to ignore cert warnings (sets option CERT_NONE)
            ssl_context = ssl.SSLContext()

            # Flag. To execute code in 'finally' block
            correct_execution = False
            exception_msg = ''

            # Execute REST call and parse data
            try:
                with urlopen(my_http_request, timeout=self.request_timeout, context=ssl_context) as u:
                    http_response = u.read()
                    # debug: print(f"REQUEST STATUS: {u.status} - REASON: {u.reason} ")

                    # Save REST response (with original IBM logic/format/syntax)
                    raw_rest_reply = json.loads(http_response)
                    self.raw_result.append(raw_rest_reply)

                    # Parse raw data (sets self.parsed_result)
                    self.parse_raw_data(raw_rest_reply, tsm_server, tsm_command)

                correct_execution = True

            # HTTP exceptions
            except urllib.error.HTTPError as e:
                if e.code == 404:
                    exception_msg = (
                        f"ERROR: Page not found! HTML RC {e.code}\n"
                        f"       Validate that the REST API is enabled and accessible\n"
                        f"       Check the settings menu in the OC at {self.base_url} with user '{tsm_user}\n"
                    )
                elif e.code == 403:
                    exception_msg = f"ERROR: Access denied! HTML RC {e.code}"
                elif e.code == 401:
                    exception_msg = (
                        f"ERROR: API Access refused! : {e.reason} HTML RC {e.code}\n"
                        f"       Try to access the operations center at {self.base_url} with user '{tsm_user}'\n"
                        f"       Validate that the REST API is enabled in the operations center settings menu\n"
                    )
                elif e.code == 500:
                    # Catches incorrect syntax AND permissions. Cannot differentiate
                    # It also catches a TSM server name that is not managed/configured on TSM OC
                    # Hopefully IBM fixes this in future, or maybe I'm blind.
                    exception_msg = "ERROR: Problem with syntax of TSM command or TSM account privileges"
                else:
                    exception_msg = f"ERROR: Incorrect request. Response: {e.code} : {e.reason}"

            # Connectivity exceptions (URLError is superclass of HTTPError that why it comes after HTTPError)
            except urllib.error.URLError as e:
                if isinstance(e.reason, socket.timeout):
                    exception_msg = f"ERROR: Timeout. Is {self.oc_address} reachable?"
                elif isinstance(e.reason, ConnectionRefusedError):
                    exception_msg = f"ERROR: Connection refused. Is {self.oc_address} port {self.oc_port} reachable?"
                else:
                    exception_msg = f"ERROR: Request error: {e.reason}"
                    # debug: print(repr(e.args[0]))

            except socket.timeout:
                exception_msg = f"ERROR: Timeout exceeded: {self.request_timeout} secs. Is {self.oc_address} reachable?"

            # If data structure is not as expected we raise ValueError in parse_raw_data()
            except ValueError as e:
                exception_msg = f"ERROR: {e.args[0]}"

            # All other exceptions
            except Exception as e:
                exception_msg = f"ERROR: Unhandled exception - Contact the developer. Text: {e} Error Type: {type(e)}"

            # Always executed, error or not
            finally:
                if correct_execution:
                    print(f" -> {tsm_server} : OK")
                else:
                    # Exception was raised, add informational record to data set
                    empty_record = {'TSM SERVER': tsm_server + ' - ' + exception_msg}
                    if 'items' in self.parsed_result:
                        self.parsed_result['items'].append(empty_record)  # Pycharm TypeChecker (ok in VSC, ignore)
                    else:
                        self.parsed_result['items'] = [empty_record]

                    print(f" -> {tsm_server} : {exception_msg}")

                # Save command and TSM server
                self.parsed_result['cmd'] = tsm_command
                self.parsed_result['tsm_srv'] = tsm_servers

            # End of Try/Catch

        # End of loop 'for every tsm server in list'

        # If there are 'items' and no 'hdr' (columns) it means we had 'No Match Found' across all TSM servers
        # So we add 'hdr' to the result, with just 1 'hdr': TSM SERVER
        if 'items' in self.parsed_result and 'hdr' not in self.parsed_result:
            self.parsed_result['hdr'] = ['TSM SERVER']

    # End of function run_command()

    @staticmethod
    def fix_value(dict_value):
        # dict_value can be a list, dict, string, int. Function returns a clean one dimensional value
        # Example: [{'val': 50}, {'val': {'def': 'GB'}}] returns '50 GB'

        # Flag to check if we recognize the pattern of 'dict_value'
        fixed = False

        if isinstance(dict_value, dict):

            # {'def': 'No', 'id': '23402'} becomes 'No'
            if 'def' in dict_value:
                dict_value = dict_value['def']
                fixed = True

            # {'secs': 1200, 'type': 0, 'tzo': 3600} becomes '1200 secs'
            elif 'secs' in dict_value:
                dict_value = datetime.datetime.fromtimestamp(dict_value['secs'])
                fixed = True

        elif isinstance(dict_value, list):
            # [] becomes '-'
            if len(dict_value) == 0:
                dict_value = '-'
                fixed = True

            elif len(dict_value) == 1:
                if isinstance(dict_value[0], dict):
                    if 'val' in dict_value[0]:

                        # [{'val': {'def': 'System', 'id': '23440'}}] becomes 'System'
                        if isinstance(dict_value[0]['val'], dict):
                            if 'def' in dict_value[0]['val']:
                                dict_value = dict_value[0]['val']['def']
                                fixed = True

                        # [{'val': 'HOSTNAME'}] becomes 'HOSTNAME'
                        elif not isinstance(dict_value[0]['val'], (dict, list)):
                            dict_value = dict_value[0]['val']
                            fixed = True

                # ['System'] becomes 'System' (one item list)
                elif not isinstance(dict_value[0], (dict, list)):
                    dict_value = dict_value[0]
                    fixed = True

            # [{'val': 1}, {'val': {'def': 'GB'}}] becomes '1 GB'
            elif len(dict_value) == 2:
                if isinstance(dict_value[0], dict) and isinstance(dict_value[1], dict):
                    if 'val' in dict_value[0] and 'val' in dict_value[1]:
                        if not isinstance(dict_value[0]['val'], (dict, list)):
                            if isinstance(dict_value[1]['val'], dict):
                                if 'def' in dict_value[1]['val']:
                                    val = dict_value[0]['val']
                                    metric = dict_value[1]['val']['def']
                                    dict_value = f"{val} {metric}"
                                    fixed = True

        else:
            fixed = True  # 'dict_value' is not a list or dict, so we leave it as-is

        if not fixed:
            dict_value = 'UNEXPECTED DATA - CONTACT DEVELOPER'

        return dict_value

    # End of function fix_value

    def parse_raw_data(self, raw_data, tsm_server, tsm_command):
        # Interpret and parse the raw REST response, save result to self.parsed_data

        data_structure_valid = False

        # A valid response is 'a list with one item: a new list. In that inner list should be at least one dict'
        if isinstance(raw_data, list) and len(raw_data) == 1:
            if isinstance(raw_data[0], list) and len(raw_data[0]) >= 1:
                if isinstance(raw_data[0][0], dict):
                    if len(raw_data[0][0]) == 1:
                        if 'msg' in raw_data[0][0]:
                            data_structure_valid = True
                    if len(raw_data[0][0]) >= 2:
                        if 'hdr' in raw_data[0][0] and 'items' in raw_data[0][0]:
                            data_structure_valid = True

            elif isinstance(raw_data[0], list) and len(raw_data[0]) == 0:  # [[]]
                # If you perform a SQL select query that returns 0 rows (no match found) then the REST API
                # sends an informative message back. But not on SP 8.1.8 and older versions. There the API returns an
                # 'empty list inside a list [[]]'. The next two lines mimic the behavior of the 'fix' in SP 8.1.9
                raw_data[0].append({'msg': {'n': '2034'}})
                data_structure_valid = True
                # If you run SP 8.1.9 or above you can remove the previous two lines, and uncomment the next line:
                # raise ValueError('Unsupported Data structure: [[]] - Invalidating')

            else:
                pass

        # Data structure is not as expected
        if not data_structure_valid:
            raise ValueError("Unsupported Data structure: Not able to validate structure")

        # Data structure valid, start parsing
        clean_data = {}

        # If key 'items' is in REST response, then we have real data (rows/columns) in the result
        # Meaning you probably executed a TSM 'query' or 'select' command
        if 'items' in raw_data[0][0]:
            # If 'hdr'[0] is a dict then it contains colID and colName, and actual data contains colID, not colName
            if not isinstance(raw_data[0][0]['hdr'][0], (dict, list)):
                for item in raw_data[0][0]['items']:
                    for header in raw_data[0][0]['hdr']:
                        item[header] = self.fix_value(item[header])

                clean_data = {'hdr': raw_data[0][0]['hdr'], 'items': raw_data[0][0]['items']}

            # 'hdr'[0] is not a dict then 'hdr' is just an ordered list with the colNames (select query)
            else:
                for item in raw_data[0][0]['items']:
                    for header in raw_data[0][0]['hdr']:
                        item[header['id']] = self.fix_value(item[header['id']])

                        # Change column ID to column title
                        item[header['def']] = item.pop(header['id'], "Not Valid")

                clean_data = {'hdr': list(item['def'] for item in raw_data[0][0]['hdr']),
                              'items': raw_data[0][0]['items']}

            # Add "TSM server" col name to header list (useful for multi-TSM server reports)
            clean_data['hdr'].insert(0, "TSM SERVER")

            # Add "TSM server":"value" to each dictionary in 'items' list (useful for multi-TSM server reports)
            for item in clean_data['items']:
                item.update({"TSM SERVER": tsm_server})

        # If there is a TSM info/warning/error message in the REST response, then we save it
        if 'msg' in raw_data[0][0]:
            raw_data[0][0]['msg'].update({"srv": tsm_server})
            raw_data[0][0]['msg'].update({"cmd": tsm_command})
            clean_data['msg'] = [raw_data[0][0]['msg']]

        # Some commands also have a message in the second item of the result list [1]
        if len(raw_data[0]) > 1 and 'msg' in raw_data[0][1]:
            if 'msg' in clean_data:
                raw_data[0][1]['msg'].update({"srv": tsm_server})
                raw_data[0][1]['msg'].update({"cmd": tsm_command})
                clean_data['msg'].append(raw_data[0][1]['msg'])
            else:
                raw_data[0][1]['msg'].update({"srv": tsm_server})
                raw_data[0][1]['msg'].update({"cmd": tsm_command})
                clean_data['msg'] = [raw_data[0][1]['msg']]

        # Data (clean_data) is now ready, got its final structure

        # Move 'items' from clean_data into instance variable parsed_result
        # This is to support command execution on multiple TSM servers
        if 'items' in clean_data:
            for data_dict in clean_data['items']:
                if 'items' in self.parsed_result:
                    self.parsed_result['items'].append(data_dict)
                else:
                    self.parsed_result['items'] = [data_dict]
            if 'hdr' not in self.parsed_result:
                self.parsed_result['hdr'] = clean_data['hdr']

        # Move 'msg' from clean_Data into instance variable parsed_result
        if 'msg' in clean_data:
            if 'msg' not in self.parsed_result:
                self.parsed_result['msg'] = []

            for m in clean_data['msg']:
                self.parsed_result['msg'].append(m)

            # If the result of the command is 'no match found' then we create an empty key/value in 'items'
            # To leave a trace behind that we did query the TSM server.
            if 'items' not in clean_data:
                for msg in clean_data['msg']:
                    if msg['n'] == '2034':  # 'No Match Found'
                        empty_record = {'TSM SERVER': tsm_server + ' - NO MATCH FOUND'}
                        if 'items' in self.parsed_result:
                            self.parsed_result['items'].append(empty_record)
                        else:
                            self.parsed_result['items'] = [empty_record]
                    else:
                        error_record = {'TSM SERVER': tsm_server + ' - ' + msg['def']}
                        if 'items' in self.parsed_result:
                            self.parsed_result['items'].append(error_record)
                        else:
                            self.parsed_result['items'] = [error_record]

    # End of function

    # Export self.parsed_result to xlsx, csv, html
    def create_report(self, report_type, file_name=None, sheet_name='Report', sheet_tab_color='FFFFFF'):

        # Validate report type parameter
        valid_report_type = {'XLSX', 'HTML', 'CSV'}
        if report_type not in valid_report_type or not isinstance(report_type, str):
            print(f"create_report(): Unknown Report type. Your options: {valid_report_type}")
            exit(1)

        # Validate target directory to store report exists
        if not os.path.exists(os.path.abspath(os.path.dirname(file_name))):
            print("create_report(): Report directory does not exist")
            exit(1)

        # Validate that file_name is a string with more than 3 chars
        if not isinstance(file_name, str) or (isinstance(file_name, str) and len(file_name) < 3):
            print("create_report(): Incorrect file_name passed")
            exit(1)

        # Validate we have data to print
        if 'items' not in self.parsed_result:
            print("No data to print, this is not normal. All exceptions should be handled. Check the 'msg' key")
            # debug: pprint(self.parsed_result)
            exit(1)

        # Function call is ok: Arguments validated

        col_names = self.parsed_result['hdr']
        clean_data = self.parsed_result['items']

        if report_type == "XLSX":

            # Initialize wb as empty workbook
            wb = Workbook()

            # If the Excel file already exists then load it. We'll place our data in a new sheet
            if os.path.exists(file_name):
                try:
                    wb = load_workbook(file_name)
                except Exception:  # At this point we already checked if path exists
                    print(f"ERROR: Could not access the target file {file_name}")
                    exit(1)
            else:
                # Excel file does not exist yet (wb already created in first step)

                # Since this is a new file we can remove the 'default' sheet
                remove_these_sheets = ['Sheet', 'Sheet1', 'Sheet 1']
                for sheet in remove_these_sheets:
                    if sheet in wb.sheetnames:
                        del wb[sheet]

            # Define styles
            create_xls_styles(wb)

            # Create a new Excel sheet, set name and tab color
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_properties.tabColor = sheet_tab_color

            # Write headers to the first row of the sheet
            ws.append(col_names)

            # Style column titles
            ws.row_dimensions[1].height = 20

            for i in range(1, len(col_names)+1):
                current_cell = str(chr(ord('A') + (i - 1)) + '1')
                ws[current_cell].style = 'title_style'

            # Write values to cells
            row_index = 2
            for item in clean_data:
                col_index = 1
                for col in col_names:
                    if col in item:
                        if item[col] is not None:
                            ws.cell(row=row_index, column=col_index).value = item[col]
                        else:
                            # If the Value is None then we set a dash in the cell '-'
                            ws.cell(row=row_index, column=col_index).value = '-'
                            # debug: print(ws.cell(row=row_index, column=col_index).coordinate)
                    else:
                        # The column name is not present as a key in the current item from 'items'
                        ws.cell(row=row_index, column=col_index).value = '-'
                    col_index += 1
                row_index += 1

            # Enable Excel Column Filtering
            ws.auto_filter.ref = ws.dimensions

            # Auto Size columns (stackoverflow flunky hack) (not happy with it)
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value + 6

            # Save Excel file
            try:
                wb.save(file_name)
            except Exception:
                print(f"ERROR: Could not save report file {file_name}")
                exit(1)

        elif report_type == 'CSV':
            # If target CSV report file exists then it gets overwritten
            try:
                with open(file_name, 'w') as output_file:
                    dict_writer = csv.DictWriter(output_file, restval="-", fieldnames=col_names, delimiter=';')
                    dict_writer.writeheader()
                    dict_writer.writerows(clean_data)
            except Exception:
                print(f"ERROR: Could not create csv file {file_name}")
                exit(1)

        elif report_type == 'HTML':
            # If target HTML report file exists then it gets overwritten
            html = "<html>\n<head>\n\t<title>Test Report</title>\n</head>\n\n<body>\n<table>\n\n\t<tr>\n"
            for col in col_names:
                html += "\t\t<th>" + col + "</th>\n"
            html += "\t</tr>\n"
            for item in clean_data:
                html += "\t<tr>\n"
                for col in col_names:
                    if col in item:
                        html += "\t\t<td>" + str(item[col]) + "</td>\n"
                    else:
                        html += "\t\t<td>-</td>\n"
                html += "\t</tr>\n"
            html += "</table>\n</body>\n</html>"

            # Save the report
            try:
                f = open(file_name, 'w')
                f.write(html)
                f.close()
            except Exception:
                print(f"ERROR: Could not create html file {file_name}")

        else:
            pass  # Unknown report type (exception already caught in beginning of function)

    # End of function create_report()

# End of class
